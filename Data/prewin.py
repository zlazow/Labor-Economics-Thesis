from openpyxl import load_workbook
from openpyxl import Workbook

wb = load_workbook("pgaData.xlsx", read_only = True) 
wb2 = Workbook()

ws = wb["Sheet1"]
ws2 = wb2.active

"""
Read each players name (going down column A). If tournament == [] check players []

If player in tournament winners list -- pt_superstar = 1. 
"""
a = ws2.cell(row = 1, column = 1)
a.value =  "pt_superstar"

r = 1
for i in ws.iter_rows('A2:B17493'):
	p, t = i

	r +=1 
	a = ws2.cell(row = r, column = 1)		

	
	if p.value in ["len mattiace", "jim furyk", "bob estes", "chris smith"] and t.value == "advil westtern open_2002":
		a.value = 1
	
	elif p.value in ["j.p. hayes", "rich beem", "chris riley", "craig parry", "rich beem"] and t.value == "air canada championship_2002":
		a.value = 1
	
	elif p.value in ["len mattiace", "jerry kelly", "jeff sluman"] and t.value == "b.c. open_2002":
		a.value = 1	
	
	elif p.value in ["len mattiace", "ian legatt", "kevin sutherland", "matt kuchar", "jose maria olazabal"] and t.value == "bay hill invitational_2002":
		a.value = 1	
	elif p.value in ["gene sauers", "rich beem", "chris riley", "craig parry"] and t.value == "bell canadian open_2002":
		a.value = 1	
	elif p.value in ["craig perks", "kevin sutherland", "matt kuchar"] and t.value == "bellsouth classic_2002":
		a.value = 1			
	elif p.value in ["charles howell iii", "phil tataurangi", "bob burns", "k.j. choi", "loren roberts"] and t.value == "buick challenge_2002":
		a.value = 1	

	elif p.value in ["shigeki maruyama", "nick price", "jim furyk", "bob estes", "rocco mediate", "k.j. choi"] and t.value == "buick classic_2002":
		a.value = 1	

	elif p.value in ["sergio garcia", "jerry kelly", "chris dimarco", "matt gogel"] and t.value == "buick invitational_2002":
		a.value = 1			
	elif p.value in ["j.p. hayes", "rich beem", "jerry kelly", "jeff sluman", "spike mcroy"] and t.value == "buick open_2002":
		a.value = 1		
	elif p.value in ["shigeki maruyama", "nick price", "jim furyk", "bob estes", "chris smith"] and t.value == "canon greater hartford open_2002":
		a.value = 1	
	elif p.value in ["craig perks", "justin leonard", "rocco mediate"] and t.value == "compaq classic_2002":
		a.value = 1
	elif p.value in ["nick price", "jim furyk", "bob estes", "chris smith"] and t.value == "fedex st. jude classic_2002":
		a.value = 1		
	elif p.value in ["len mattiace", "ian legatt", "kevin sutherland", "chris dimarco", "matt gogel", "jose maria olazabal"] and t.value == "genuity championship_2002":
		a.value = 1		
	elif p.value in ["craig perks", "justin leonard"] and t.value == "greater greensboro chrysler classic_2002":
		a.value = 1												
	elif p.value in ["len mattiace", "jerry kelly", "bob estes", "chris smith"] and t.value == "greater milwaukee open_2002":
		a.value = 1				
	elif p.value in ["len mattiace", "ian legatt", "kevin sutherland", "matt gogel", "jose maria olazabal"] and t.value == "honda classic_2002":
		a.value = 1				
	elif p.value in ["charles howell iii", "john rollins", "dan forsman", "k.j. choi", "loren roberts"] and t.value == "invensys classic at las vegas_2002":
		a.value = 1					
	elif p.value in ["len mattiace", "jerry kelly", "jeff sluman", "spike mcroy"] and t.value == "john deere classic_2002":
		a.value = 1	

	elif p.value in ["shigeki maruyama", "nick price", "jim furyk", "justin leonard", "rocco mediate", "k.j. choi"] and t.value == "kemper insurance open_2002":
		a.value = 1	

	elif p.value in ["shigeki maruyama", "justin leonard", "rocco mediate", "k.j. choi"] and t.value == "mastercard colonial_2002":
		a.value = 1	

	elif p.value in ["shigeki maruyama", "nick price", "justin leonard", "rocco mediate", "k.j. choi"] and t.value == "memorial tournament_2002":
		a.value = 1	
	elif p.value in ["gene sauers", "john rollins", "dan forsman", "k.j. choi", "loren roberts"] and t.value == "michelob championship at kingsmill_2002":
		a.value = 1	

	elif p.value in ["j.p. hayes", "rich beem", "chris riley", "spike mcroy"] and t.value == "nec invitational_2002":
		a.value = 1	
	elif p.value in ["sergio garcia", "jerry kelly", "chris dimarco", "matt gogel", "jose maria olazabal"] and t.value == "nissan open_2002":
		a.value = 1	
	elif p.value in ["sergio garcia", "jerry kelly"] and t.value == "phoenix open_2002":
		a.value = 1
	elif p.value in ["j.p. hayes", "rich beem", "spike mcroy"] and t.value == "reno-tahoe open_2002":
		a.value = 1		

	elif p.value in ["gene sauers", "john rollins", "rich beem", "chris riley", "craig parry"] and t.value == "sei pennsylvania classic_2002":
		a.value = 1		

	elif p.value in ["craig perks", "ian legatt", "kevin sutherland", "matt kuchar"] and t.value == "shell houston open_2002":
		a.value = 1		

	elif p.value in ["sergio garcia"] and t.value == "sony open_2002":
		a.value = 1		
	elif p.value in ["charles howell iii", "phil tataurangi", "bob burns", "jonathan byrd", "loren roberts"] and t.value == "southern farm bureau classic_2002":
		a.value = 1		
	elif p.value in ["gene sauers", "john rollins", "dan forsman", "rich beem", "Chris Riley", "craig parry"] and t.value == "tampa bay classic_2002":
		a.value = 1		

	elif p.value in ["len mattiace", "jerry kelly", "chris dimarco", "matt gogel", "jose maria olazabal"] and t.value == "touchstone energy tucson open_2002":
		a.value = 1		

	elif p.value in ["gene sauers", "john rollins", "dan forsman", "k.j. choi", "craig parry"] and t.value == "valero texas open_2002":
		a.value = 1		
	elif p.value in ["craig perks", "justin leonard", "rocco mediate", "k.j. choi"] and t.value == "verizon byron nelson classic_2002":
		a.value = 1		

	elif p.value in ["craig perks", "matt kuchar"] and t.value == "worldcom classic_2002":
		a.value = 1		

	elif p.value in ["brad faxon", "olin browne", "mark calcavecchia", "jason bohn"] and t.value == "84 lumber classic_2005":
		a.value = 1		

	elif p.value in ["michael campbell", "padraig harrington", "jim furyk", "sean o'hair", "sergio garcia"] and t.value == "b.c. open_2005":
		a.value = 1		

	elif p.value in ["ted purdy", "peter leonard", "tim petrovic"] and t.value == "bank of america colonial_2005":
		a.value = 1		

	elif p.value in ["michael campbell", "ted purdy", "kenny perry", "justin leonard", "bart bryant", "sergio garcia"] and t.value == "barclays classic_2005":
		a.value = 1	
	elif p.value in ["adam scott", "david toms", "geoff ogilvy", "padraig harrington"] and t.value == "bay hill invitational_2005":
		a.value = 1		

	elif p.value in ["brad faxon", "olin browne", "jason bohn"] and t.value == "bell canadian open_2005":
		a.value = 1		
	elif p.value in ["fred funk", "david toms", "geoff ogilvy", "padraig harrington", "kenny perry"] and t.value == "bellsouth classic_2005":
		a.value = 1		
	elif p.value in ["ted purdy", "kenny perry", "justin leonard", "bart bryant", "tim petrovic"] and t.value == "booz allen classic_2005":
		a.value = 1

	elif p.value in ["ben crane", "vaugh taylor"] and t.value == "buick championship_2005":
		a.value = 1			
	elif p.value in ["fred funk", "padraig harrington", "kenny perry"] and t.value == "mci heritage_2005":
		a.value = 1			
	elif p.value in ["ben crane", "padraig harrington", "jim furyk", "sean o'hair", "jason bohn"] and t.value == "buick open_2005":
		a.value = 1			
	elif p.value in ["west short", "lucas glover", "jason gore", "robert gamez", "k.j. choi"] and t.value == "chrysler championship_2005":
		a.value = 1			
	elif p.value in ["brad faxon", "olin browne", "mark calcavecchia", "jason gore", "robert gamez", "jason bohn"] and t.value == "chrysler classic of greensboro_2005":
		a.value = 1			
	elif p.value in ["adam scott", "david toms", "justin leonard"] and t.value == "chrysler classic of tucson_20052":
		a.value = 1			
	elif p.value in ["michael campbell", "padraig harrington", "kenny perry", "justin leonard", "bart bryant", "sergio garcia"] and t.value == "cialis westtern open_2005":
		a.value = 1				
	elif p.value in ["brad faxon", "jason bohn"] and t.value == "deutsche bank championship_2005":
		a.value = 1	
	elif p.value in ["ted purdy", "kenny perry", "peter leonard", "tim petrovic"] and t.value == "fedex st. jude classic_2005":
		a.value = 1	
	elif p.value in ["adam scott", "david toms", "geoff ogilvy", "justin leonard"] and t.value == "ford championship at doral_2005":
		a.value = 1	
	elif p.value in ["adam scott", "david toms", "geoff ogilvy"] and t.value == "honda classic_2005":
		a.value = 1	
	elif p.value in ["michael campbell", "padraig harrington", "jim furyk", "justin leonard", "bart bryant", "sergio garcia"] and t.value == "john deere classic_2005":
		a.value = 1	
	elif p.value in ["ted purdy", "kenny perry", "justin leonard", "tim petrovic"] and t.value == "memorial tournament_2005":
		a.value = 1	
	elif p.value in ["stuart appleby", "justin leonard"] and t.value == "nissan open_2005":
		a.value = 1	
	elif p.value in ["ben crane", "jason bohn"] and t.value == "reno-tahoe open_2005":
		a.value = 1	
	elif p.value in ["fred funk", "peter leonard", "padraig harrington", "kenny perry"] and t.value == "shell houston open_2005":
		a.value = 1	
	elif p.value in ["stuart appleby"] and t.value == "sony open in hawaii_2005":
		a.value = 1	
	elif p.value in ["michael campbell", "padraig harrington", "jim furyk", "sean o'hair", "jason bohn"] and t.value == "u.s. bank championship in milwaukee_2005":
		a.value = 1	
	elif p.value in ["fred funk", "peter leonard", "tim petrovic"] and t.value == "wachovia championship_2005":
		a.value = 1	
	elif p.value in ["fred funk", "peter leonard", "kenny perry"] and t.value == "zurich classic of new orleans_2005":
		a.value = 1	
	elif p.value in ["jim furyk", "brett wetterich", "aaron baddeley", "stuart appleby", "chris couch"] and t.value == "bank of america colonial_2006":
		a.value = 1	
	elif p.value in ["brad faxon", "olin browne", "mark calcavecchia", "jason gore", "jason bohn"] and t.value == "valero texas open_2005":
		a.value = 1	
	elif p.value in ["jim furyk", "brett wetterich", "tim herron", "jeff maggert", "carl pettersson", "chris couch"] and t.value == "barclays classic_2006":
		a.value = 1	
	elif p.value in ["rory sabbatini", "geoff ogilvy", "kirk triplett", "luke donald", "arron oberholser"] and t.value == "bay hill invitational_2006":
		a.value = 1	
	elif p.value in ["stephen ames", "geoff ogilvy", "kirk triplett", "luke donald", "rod pampling"] and t.value == "bellsouth classic_2006":
		a.value = 1	
	elif p.value in ["geoff ogilvy", "brett wetterich", "tim herron", "jeff maggert", "carl pettersson"] and t.value == "booz allen classic_2006":
		a.value = 1	
	elif p.value in ["rory sabbatini", "geoff ogilvy", "chad campbell", "j.b. holmes", "arron oberholser"] and t.value == "chrysler classic of tucson_2006":
		a.value = 1	
	elif p.value in ["jim furyk", "brett wetterich", "tim herron", "aaron baddeley", "stuart appleby", "chris couch"] and t.value == "fedex st. jude classic_2006":
		a.value = 1	
	elif p.value in ["rory sabbatini", "geoff ogilvy", "kirk triplett", "j.b. holmes", "arron oberholser"] and t.value == "ford championship at doral_2006":
		a.value = 1	
	elif p.value in ["rory sabbatini", "geoff ogilvy", "kirk triplett", "j.b. holmes", "arron oberholser"] and t.value == "honda classic_2006":
		a.value = 1	
	elif p.value in ["jim furyk", "brett wetterich", "tim herron", "jeff maggert", "stuart appleby", "chris couch"] and t.value == "memorial tournament_2006":
		a.value = 1	
	elif p.value in ["stuart appleby", "david toms", "chad campbell", "j.b. holmes", "arron oberholser"] and t.value == "nissan open_2006":
		a.value = 1	
	elif p.value in ["stephen ames", "aaron baddeley", "luke donald", "rod pampling"] and t.value == "shell houston open_2006":
		a.value = 1	


	elif p.value in ["stuart appleby"] and t.value == "sony open in hawaii_2006":
		a.value = 1	
	elif p.value in ["stephen ames", "luke donald", "rod pampling"] and t.value == "verizon heritage_2006":
		a.value = 1	
	elif p.value in ["stephen ames", "aaron baddeley", "stuart appleby", "chris couch"] and t.value == "wachovia championship_2006":
		a.value = 1	
	elif p.value in ["stephen ames", "aaron baddeley", "stuart appleby", "rod pampling"] and t.value == "zurich classic of new orleans_2006":
		a.value = 1	
	else:
		a.value = 0

wb2.save("bambalam.xlsx")